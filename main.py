import pandas as pd
import datetime
import openpyxl
import csv
import chardet

class mesToExcel:
    file_path = r"C:\TEMP\file.csv"
    optimized_file = r"C:\TEMP\betterFile.csv"

    my_first_file = r"C:\TEMP\normalize_csv_01.csv"
    # Read the file with the detected encoding
    with open(file_path, 'r', encoding='UTF-16') as input_file, \
            open(optimized_file, 'w', newline='', encoding='UTF-8') as output_file:
        csv_reader = csv.reader(input_file, delimiter='\t')
        csv_writer = csv.writer(output_file, delimiter=';')

        row_num = 0
        for row in csv_reader:
            if row_num < 1:
                # skip first two rows
                row_num += 1
                continue

            cleaned_row = []
            for cell in row:
                if '\t' in cell:
                    split_cell = cell.split('\t')
                    cleaned_cell = ';'.join([part.strip() for part in split_cell])
                    cleaned_row.append(cleaned_cell)
                else:
                    cleaned_row.append(cell.strip())
            csv_writer.writerow(cleaned_row)




    df = pd.read_csv(optimized_file, delimiter=';')

    
    
    Rollo = [["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"]]

    GuideRails = [["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"]]

    Assembly = [["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"],
                ["xxxxx", "ccccc", "rrrrr"]]

    Encapsulation = [["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"],
                     ["xxxxx", "ccccc", "rrrrr"]]

    yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
    date = yesterday.strftime("%Y-%m-%d")

    base = [date, "rrrrr", "rrrrr", "rrrrr", "rrrrr", "rrrrr", "rrrrr", "rrrrr","rrrrr", "rrrrr", "rrrrr",
              "rrrrr", "rrrrr", "rrrrr", "rrrrr", "rrrrr", "rrrrr", "rrrrr"]


def oneExcelFile(csv, firstRow, template,fileName):
    # Create a new workbook object
    workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Write the array contents to the first row
    for col_num, value in enumerate(firstRow, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = value
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.font = openpyxl.styles.Font(size=14, bold=True)
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='808080')
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                             right=openpyxl.styles.Side(style='thin'),
                                             top=openpyxl.styles.Side(style='thin'),
                                             bottom=openpyxl.styles.Side(style='thin'))

    # Write Liberec
    sheet['A2'] = 'Liberec'   ###  sheet.cell(row=2, column=1).value = 'string'

    #Add data for one fileSheet
    for item in template:
        workcenter = item[0]
        commoncost = item[1]
        rework = item[2]

        print(workcenter, commoncost, rework)

    for index, item in enumerate(template):
        for row in csv.iterrows():
            if row[1][2] == item[0]:
                name = row[1][2]
                ok = row[1][3]
                nok = row[1][4]
                attendance = (float(row[1][5].replace(",", ".")))

                nonRepTimes = (float(row[1][7].replace(",", ".")))
                disruptions = (float(row[1][8].replace(",", ".")))
                partsxTEB = (float(row[1][12].replace(",", ".")))
                breakDuration = (float(row[1][14].replace(",", ".")))

                attendanceComCost = '0'
                attendanceRew = '0'

                newRow = index + 3


                # NAME
                cell = sheet.cell(row=newRow, column=1)
                cell.value = name
                # OK
                cell = sheet.cell(row=newRow, column=2)
                cell.value = ok
                # NOK
                cell = sheet.cell(row=newRow, column=3)
                cell.value = nok
                # FTQ
                cell = sheet.cell(row=newRow, column=4)
                cell.value = '=B'+str(newRow)+'/(B'+str(newRow)+'+C'+str(newRow)+')'
                cell.number_format = '0.00%'



                # Act / Plan %
                cell = sheet.cell(row=newRow, column=6)
                cell.value = '=(B' + str(newRow) + '+C' + str(newRow) + ')/E' + str(newRow)
                cell.number_format = '0.00%'
                # Rout time
                cell = sheet.cell(row=newRow, column=7)
                cell.value = '=(P' + str(newRow) + '*60)/(B' + str(newRow) + '+C' + str(newRow)+')'
                # PartsxTEB
                cell = sheet.cell(row=newRow, column=8)
                cell.value = partsxTEB
                # PartsxTEB%
                cell = sheet.cell(row=newRow, column=9)
                cell.value = '=H' + str(newRow) + '/(P' + str(newRow) + '-J' + str(newRow) + ')'
                cell.number_format = '0.00%'


                # RW Oper
                cell = sheet.cell(row=newRow, column=11)
                cell.value = '=J' + str(newRow) + '/7.5'
                # No Of Oper

                cell = sheet.cell(row=newRow, column=17)
                cell.value = '=P' + str(newRow) + '/7.5'


                # Com cost attendace
                for rowX in csv.iterrows():
                    if rowX[1][2] == item[1]:
                        attendanceComCost = (float(rowX[1][5].replace(",", ".")))



                # Rework
                for rowY in csv.iterrows():
                    if rowY[1][2] == item[2]:
                        attendanceRew = (float(rowY[1][5].replace(",", ".")))

                attendanceRew = float(attendanceRew)
                attendanceComCost = float(attendanceComCost)

                #RW h
                cell = sheet.cell(row=newRow, column=10)
                cell.value = attendanceRew

                #Com. Cost
                cell = sheet.cell(row=newRow, column=12)
                cell.value = attendanceComCost


                final_attendance = attendance + attendanceRew + attendanceComCost

                # attendance h
                cell = sheet.cell(row=newRow, column=16)
                cell.value = final_attendance

                PLF = (partsxTEB + breakDuration) / final_attendance
                NonrepTimes = (100 - nonRepTimes) / (final_attendance * 100)

                # PLF
                cell = sheet.cell(row=newRow, column=13)
                cell.value = PLF
                cell.number_format = '0.00%'

                # non rep times
                cell = sheet.cell(row=newRow, column=14)
                cell.value = NonrepTimes
                cell.number_format = '0.00%'

                #disruptions
                cell = sheet.cell(row=newRow, column=15)
                cell.value = disruptions / final_attendance
                cell.number_format = '0.00%'

    # Auto-fit the columns to the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.4
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save("PLF_"+fileName+".xlsx")

if __name__ == '__main__':
    oneExcelFile(mesToExcel.df, mesToExcel.base, mesToExcel.rrrrr, "rrrrr")
    oneExcelFile(mesToExcel.df, mesToExcel.base, mesToExcel.rrrrr, "rrrrr")
    oneExcelFile(mesToExcel.df, mesToExcel.base, mesToExcel.rrrrr, "rrrrr")
    oneExcelFile(mesToExcel.df, mesToExcel.base, mesToExcel.rrrrr, "rrrrr")
